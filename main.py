import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import time
import re

# CAMINHO DO CHROMEDRIVER
CHROMEDRIVER_PATH = "./chromedriver.exe"

# URL do Climatempo para São Paulo
URL = "https://www.climatempo.com.br/previsao-do-tempo/cidade/558/saopaulo-sp"

# Nome do arquivo Excel
ARQUIVO = "dados_temperatura.xlsx"

def buscar_e_salvar():
    try:
        # Configura o Selenium
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        service = Service(CHROMEDRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=options)
        driver.get(URL)
        time.sleep(3)  # Espera carregar

        # CAPTURA TEMPERATURA
        try:
            temp_element = driver.find_element(By.CLASS_NAME, "temperature")
            temperatura = temp_element.text.strip().replace("°", "")
        except Exception as e:
            temperatura = "Não encontrado"

        # CAPTURA UMIDADE (busca ampla)
        umidade = "Não encontrado"
        try:
            # Busca todos os elementos de texto visíveis (span, div, li)
            elements = driver.find_elements(By.XPATH, "//*[contains(text(), 'Umidade')]")
            for elem in elements:
                texto = elem.text
                # Pega o número após "Umidade", ex: "Umidade 67%"
                if "Umidade" in texto:
                    partes = texto.split()
                    for i, parte in enumerate(partes):
                        if "Umidade" in parte and i+1 < len(partes):
                            prox = partes[i+1]
                            numero = ''.join(filter(str.isdigit, prox))
                            if numero:
                                umidade = numero
                                break
                    if umidade != "Não encontrado":
                        break
            # Fallback: pega qualquer número com % depois da palavra Umidade
            if umidade == "Não encontrado":
                page_text = driver.find_element(By.TAG_NAME, "body").text
                match = re.search(r'Umidade[^\d]*(\d{2})%', page_text)
                if match:
                    umidade = match.group(1)
        except Exception as e:
            umidade = "Não encontrado"

        # Data e hora atuais
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        driver.quit()

        # Salva no Excel
        if os.path.exists(ARQUIVO):
            wb = load_workbook(ARQUIVO)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Data/Hora", "Temperatura (°C)", "Umidade do Ar (%)"])

        ws.append([data_hora, temperatura, umidade])
        wb.save(ARQUIVO)

        messagebox.showinfo("Sucesso!", f"Dados salvos:\n\nData/Hora: {data_hora}\nTemperatura: {temperatura} °C\nUmidade: {umidade}%")
    except Exception as e:
        messagebox.showerror("Erro!", f"Ocorreu um erro:\n{e}")

def iniciar_interface():
    janela = tk.Tk()
    janela.title("Captador de Temperatura de São Paulo")
    janela.geometry("350x200")

    label = tk.Label(janela, text="Clique para buscar temperatura e umidade do ar\nna cidade de São Paulo agora.", font=("Arial", 11))
    label.pack(pady=30)

    botao = tk.Button(janela, text="Buscar previsão", font=("Arial", 14), command=buscar_e_salvar)
    botao.pack(pady=10)

    janela.mainloop()

if __name__ == "__main__":
    iniciar_interface()
